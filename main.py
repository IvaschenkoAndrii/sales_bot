import types

import openpyxl
import telebot
import glob, os
from telebot import types
# для указание типов

bot = telebot.TeleBot('5516232740:AAFqZ1hwAKhY9p2WmWkGniZerCHt3iJOKSg')


# Стартовое сообщение и создание кнопок бота

@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_dafi = types.KeyboardButton('Днепр, Дафи')
    btn_yavornickogo = types.KeyboardButton('Днепр, Яворницкого')
    btn_dreamtown = types.KeyboardButton('Киев, Дримтаун')
    btn_ruska = types.KeyboardButton('Львов, Руська')
    markup.add(btn_dafi, btn_yavornickogo, btn_dreamtown, btn_ruska)
    bot.send_message(message.chat.id, text="Добрый день! Выберите магазин из списка!".format(message.from_user),
                     reply_markup=markup)


@bot.message_handler(content_types=['text'])
def func(message):
    file = glob.glob("D:\Python\*.xlsx")[0]

    # читаем excel-файл
    wb = openpyxl.load_workbook(file)
    # получаем активный лист
    sheet = wb.active

    # уставливаем значения продаж изначально в 0
    sales_dream = 0
    sales_dafi = 0
    sales_yavornickogo = 0
    sales_ruska = 0

    # Перебираем значения строк, ищем названия магазинов и присваиваем переменной сумму продаж на текущий момент

    for i in range(10, 14):

        if sheet.cell(row=i, column=1).value == 'Киев ТРЦ Дрим Таун':
            sales_dream = sheet.cell(row=i, column=6).value


        elif sheet.cell(row=i, column=1).value == 'Днепр ТРЦ Дафи':
            sales_dafi = sheet.cell(row=i, column=6).value


        elif sheet.cell(row=i, column=1).value == 'Днепр пр д Яворницкого 46':
            sales_yavornickogo = sheet.cell(row=i, column=6).value

        elif sheet.cell(row=i, column=1).value == 'Львов, Руська 12':
            sales_ruska = sheet.cell(row=i, column=6).value

    if (message.text == "Днепр, Яворницкого"):
        bot.send_message(message.chat.id, 'Явроницкого продали на ' + str(sales_yavornickogo) + ' грн')

    elif (message.text == "Днепр, Дафи"):
        bot.send_message(message.chat.id, 'Дафи продали на ' + str(sales_dafi) + ' грн')

    elif (message.text == "Киев, Дримтаун"):
        bot.send_message(message.chat.id, 'Дримтаун продали на ' + str(sales_dream) + ' грн')

    elif (message.text == "Львов, Руська"):
        bot.send_message(message.chat.id, 'Руська продали на ' + str(sales_ruska) + ' грн')

    else:
        bot.send_message(message.chat.id, 'Нет такого магазина!!!')


bot.polling(none_stop=True)

